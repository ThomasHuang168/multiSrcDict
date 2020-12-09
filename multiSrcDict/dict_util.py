from openpyxl import Workbook

# ###########################################################################
# ############################# Iterate Scripts #############################
# ###########################################################################
# from docx import Document
# import os
# def process_script(docx_list, docx_path, kernel, end_file_callback=None, start_file_callback=None):
#     # this function iterate scripts (in docx) listed in docx_list and localed in docx_path
#     # and then read paragraphs in scripts and pass them into kernel and run
#     with open(docx_list, 'r', encoding="utf-8") as fr:
#         for line in fr:
#             doc_name = line.strip()
#             doc_path = os.path.join(docx_path, doc_name)
#             doc = Document(doc_path)

#             if start_file_callback is not None:
#                 start_file_callback(doc_name)

#             for para in doc.paragraphs:
#                 text = para.text
#                 kernel(text)

#             if end_file_callback is not None:
#                 end_file_callback(doc_name)


###########################################################################
############################ Export Dictionary ############################
###########################################################################

# def dict2xlsx(dict_, xlsx_file, filter=None, existingWorkSheet=None):
#     if type(dict_) != dict and type(dict_) != list:
#         raise TypeError("dict should be of type dict or list")
    
#     if existingWorkSheet:
#         ws  = existingWorkSheet
#     else:
#         wb = Workbook()
#         ws = wb.active
#     index = 1
#     key_val = list()
#     if type(dict_) == dict:
#         key_val = dict_.items()
#     else:
#         key_val = dict_
#     if None is filter:
#         for key, value in key_val:
#             ws['A{}'.format(index)] = key
#             ws['B{}'.format(index)] = value
#             index = index + 1
#     else:
#         for key, value in key_val:
#             if filter(key,value):
#                 ws['A{}'.format(index)] = key
#                 ws['B{}'.format(index)] = value
#                 index = index + 1
#     if not existingWorkSheet:
#         if len(xlsx_file) > 0:
#             wb.save(xlsx_file)
#         return wb

def dict2file(dict_, file, filter=None, encoding="utf-8"):
    if file.endswith(".txt"):
        key_val = list()
        if type(dict_) == dict:
            key_val = dict_.items()
        else:
            key_val = dict_
        with open(file, 'w', encoding=encoding) as fw:
            # for word, count in word_count.items():
            if None is filter:
                for key, value in key_val:
                    fw.write(key + '\t' + str(value) + '\n')
            else:
                for key, value in key_val:
                    if filter(key, value):
                        fw.write(key + '\t' + str(value) + '\n')
    elif file.endswith(".xlsx"):
        dict2xlsx(dict_, file, filter)
    else:
        print("formats other than txt and xlsx are not supported yet")

###########################################################################
############################# Load Dictionary #############################
###########################################################################

def load_dict(dict_file, line2KeyValueList, redundunt_file=None, encoding="utf-8", start_from_line=0):
    _dict = dict()
    _redu = dict()
    with open(dict_file, 'r', encoding=encoding) as df:
        current_line_number = 0
        for line in df:
            if current_line_number < start_from_line:
                current_line_number = current_line_number + 1
                continue
            key_value = line2KeyValueList(line)
            if key_value[0] not in _dict.keys():
                _dict[key_value[0]] = key_value[1]
            elif key_value[0] not in _redu.keys():
                _redu[key_value[0]] = 1
            else:
                _redu[key_value[0]] = _redu[key_value[0]] + 1
        if redundunt_file != None:
            dict2file(_redu, redundunt_file)
    return _dict


# ~~~~~~~ Examples of line2KeyValue ~~~~~~~ #
import re

def line2KeyValue_space(line):
    word_phone = re.split(" ", line.strip())
    return word_phone

def line2KeyValue_tab(line):
    word_phone = re.split("\t", line.strip())
    # word_phone[0].decode("utf8").encode("big5")
    return word_phone

def line2KeyValue_tab_before_slash(line):
    word_phone = re.split("\t", line.strip())
    if len(re.split("/", word_phone[1].strip())):
        word_phone[1] = re.split("/", word_phone[1].strip())[0]
    # word_phone[0].decode("utf8").encode("big5")
    return word_phone

def line2KeyValue_tab_before_space(line):
    word_phone = re.split("\t", line.strip())
    if len(re.split(" ", word_phone[1].strip())):
        word_phone[1] = re.split(" ", word_phone[1].strip())[0]
    # word_phone[0].decode("utf8").encode("big5")
    return word_phone


###########################################################################
############################### Export List ###############################
###########################################################################

def list2file(list_, file, filter=None, encoding="utf-8"):
    if file.endswith(".txt"):
        with open(file, 'w', encoding=encoding) as fw:
            numItem = len(list_)
            if None is filter:
                for index, item in enumerate(list_, 1):
                    fw.write(item)
                    if index < numItem:
                        fw.write('\n')
            else:
                for index, item in enumerate(list_, 1):
                    if filter(index, item):
                        fw.write(item)
                        if index < numItem:
                            fw.write('\n')
    else:
        print("formats other than txt and xlsx are not supported yet")

###########################################################################
########################### Counting Dictionary ###########################
###########################################################################
import operator
class countDict:
    def __init__(self, endFile_no_ext, savetxt=True, savexlsx=True):
        self.endFile_no_ext = endFile_no_ext
        self.savetxt = savetxt
        self.savexlsx = savexlsx
        self.dict_ = dict()

    def add(self, item2count):
        self.add_by_weight(item2count, 1)

    def add_by_weight(self, item2count, item_weight):
        """
        20200323 used to count the frquency by weight
        """
        if item2count not in self.dict_.keys():
            self.dict_[item2count] = item_weight
        else:
            self.dict_[item2count] += item_weight

    def export(self, filter=None, existingWorkSheet=None):
        if len(self.dict_) > 0:
            dict_sorted = sorted(self.dict_.items(), key=operator.itemgetter(1), reverse=True)
            if self.savetxt:
                dict2file(dict_sorted, self.endFile_no_ext+".txt", filter=filter)
            if self.savexlsx:
                dict2file(dict_sorted, self.endFile_no_ext+".xlsx", filter=filter)
            if existingWorkSheet:
                dict2xlsx(dict_sorted, "", filter=filter, existingWorkSheet=existingWorkSheet)

###########################################################################
#################### fetch def from Multiple Dictionary ###################
###########################################################################
import collections
class MultiDict(collections.MutableMapping):
    """A dictionary fetched definition from mutliple dictionaries by user-defined priority function, and estimate the statistics of dictionary usage"""
    # inspired by https://stackoverflow.com/questions/3387691/how-to-perfectly-override-a-dict
    def __init__(self, *args, **kwargs):
        self.dicts = dict()
        self.store = dict()
        self.update(dict(*args, **kwargs))
        self.dictPriorityAndMethod = dict()

    def __getitem__(self, dict_key):
        return self.get_def_and_update_count(dict_key, 1)

    def __setitem__(self, dict_name, dict_content):
        # ~~~~~~~~~~~ append dictionary ~~~~~~~~~~~ #
        if type(dict_content) is dict:
            self.dicts[self.__keytransform__(dict_name)] = dict_content

    def __delitem__(self, key):
        del self.store[self.__keytransform__(key)]

    def __iter__(self):
        return iter(self.store)

    def __len__(self):
        return len(self.store)

    def __keytransform__(self, key):
        return key

    def setDicts(self, dicts):
        if type(dicts) == dict:
            self.dicts = dicts
        else:
            raise TypeError("dicts ({}) should be dict".format(dicts))

    def setPriority(self, dictPriorityAndMethod):
        if type(dictPriorityAndMethod) == list:
            # self.dictPriority = dictPriorityAndMethod
            self.dictPriorityAndMethod = self.dictPriorityAndMethod.fromkeys(set(dictPriorityAndMethod))
        elif type(dictPriorityAndMethod) == dict:
            self.dictPriorityAndMethod = dictPriorityAndMethod
        else:
            raise TypeError("dictPriorityAndMethod ({}) should be list or dict".format(type(dictPriorityAndMethod)))

    def export(self, endFile_no_ext, filter=None, savetxt=True, savexlsx=True, existingWorkSheet=None):

        lim_dict = ord('Z') - ord('A')
        if len(self.dicts) > lim_dict:
            raise ValueError("please limit the number of dict < {}".format(lim_dict))
        elif len(self.dictPriorityAndMethod) == 0:
            dict_names = list(self.dicts.keys())
        elif type(self.dictPriorityAndMethod) == dict:
            dict_names = list(self.dictPriorityAndMethod.keys())
        elif type(self.dictPriorityAndMethod) == list:
            dict_names = self.dictPriorityAndMethod
        else:
            raise NotImplementedError("{} is not supported yet".format(type(self.dictPriorityAndMethod)))

        dict_names.append("not exists")

        if existingWorkSheet:
            ws  = existingWorkSheet
        else:
            wb = Workbook()
            ws = wb.active
        # ~~~~~~~~~~~~~ First Element ~~~~~~~~~~~~~ #
        indRow = 1
        indCol = ord('A')
        ws['{}{}'.format(chr(indCol),indRow)] = "Story words"

        # ~~~~~~~~~~~~~~~ First Row ~~~~~~~~~~~~~~~ #
        indCol = indCol + 1
        ws['{}{}'.format(chr(indCol),indRow)] = "Occurence"
        dictname2indCol = dict()
        for dict_name in dict_names:
            indCol = indCol + 1
            ws['{}{}'.format(chr(indCol),indRow)] = dict_name
            dictname2indCol[dict_name] = indCol
        indRow = indRow + 1

        for key, record in self.store.items():
            indCol = ord('A')
            word = key
            # ~~~~~~~ First Element in each row ~~~~~~~ #
            ws['{}{}'.format(chr(indCol),indRow)] = word

            # ~~~~ Content in each (indCol, indRow) ~~~ #
            # ~~~~~~~~~~~~~~~ Occurence ~~~~~~~~~~~~~~~ #
            indCol = indCol + 1
            ws['{}{}'.format(chr(indCol),indRow)] = record[1]

            # ~~~~~~~~~~~~~~~ definition ~~~~~~~~~~~~~~ #
            indCol = dictname2indCol[record[2]]
            def_ = record[0]
            if len(def_) == 0:
                def_ = "None"
            ws['{}{}'.format(chr(indCol),indRow)] = def_

            indRow = indRow + 1

        if not existingWorkSheet:
            wb.save(endFile_no_ext + ".xlsx")

    def get_def_and_update_count(self, dict_key, weight):
        """
        20200323 used to fetch def and count the frquency chunk by chunk
        """
        if len(self.dicts) == 0 and len(self.store) == 0:
            raise ValueError("dicts and dict_cache(store) cannot both be empty")

        # ~~~~~~~ fetch def from dictionary ~~~~~~~ #
        dictPriorityList = list()
        if len(self.dictPriorityAndMethod) > 0:
            dictPriorityList = list(self.dictPriorityAndMethod.keys())
        else:
            dictPriorityList = list(self.dicts.keys())

        dict_val = ""
        if dict_key in self.store.keys():
            # ~~~~~~~~~~~~ fetch definition ~~~~~~~~~~~ #
            dict_val = self.store[dict_key][0]
            # ~~~~~~~~~~~~ update counting ~~~~~~~~~~~~ #
            self.store[dict_key][1] += weight
        else:
            dict_entry = list()
            for dictionary in dictPriorityList:
                if len(self.dictPriorityAndMethod) > 0:
                    if dictionary in self.dictPriorityAndMethod.keys():
                        if self.dictPriorityAndMethod[dictionary] is not None:
                            search = self.dictPriorityAndMethod[dictionary]
                            # ~~~~~~~~~~ expect def if found ~~~~~~~~~~ #
                            # ~~~~~~~~ expect None if not found ~~~~~~~ #
                            dict_val = search(dict_key, self.dicts[dictionary])

                if len(dict_val) == 0 and dict_key in self.dicts[dictionary].keys():
                    dict_val = self.dicts[dictionary][dict_key]

                if len(dict_val):
                    # ~~~~~~ create entry for self.store ~~~~~~ #
                    # ~~~~~~~~~ def, count, dict_name ~~~~~~~~~ #
                    dict_entry = [dict_val, weight, dictionary]
                    break

            if len(dict_entry) == 0:
                dict_entry = ["", weight, "not exists"]
            self.store[dict_key] = dict_entry
        return dict_val


#  Examples of Method in dictPriorityAndMethod  #
def search_by_char(key_, dict_):
    if type(dict_) is not dict or type(key_) is not str:
        raise TypeError("dict_ ({}) should be dict and key_ ({}) should be str".format(type(dict_), type(key_)))

    val_ = ""
    if key_ in dict_.keys():
        val_ = dict_[key_]
    elif len(key_) > 0:
        val_c = ""
        for c in key_:
            if c not in dict_.keys():
                # ~~~~~~ return None once search fail ~~~~~ #
                val_c = ""
                break
            if len(val_c):
                val_c = val_c + "-" + dict_[c]
            else:
                val_c = val_c + dict_[c]
        if val_c != "":
            val_ = val_c
    return val_

def Test_MultiDict():
    # 20191106 
    from shortcuts import dicts
    md = MultiDict()
    md.setDicts(dicts)

    dictPriorityList = {
        "culex": None,
        "cupdict": search_by_char
    }

    md.setPriority(dictPriorityList)

    string2transcribed = "鍾意 返咗 唸諗 揾 畀 郁  擔返 返嚟 揼垃圾 掂 拎返 攰 大聲嗌 變番 瞓喺地下 鍾意 呢咗喺梳化下面 孭住 掟 耷低頭 嗌交 掹出嚟 跌咗落 打喊路 掛住你 郁咗一郁 應承 整到暈低咗 氹掂佢 瞓得好淋(nam6) 奸賴 唔使特登去 敲吓 傾偈 碌咗出去 有塵黐(痴)左係度 撳鐘"
    transcribeList = string2transcribed.split(" ")
    for w in transcribeList:
        t_w = md[w]
        if None is not t_w:
            print (t_w + "\n")

    md.export("Test_MultiDict")